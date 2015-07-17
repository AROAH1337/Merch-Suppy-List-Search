#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import sys

rowCount = 1
found = False
currentCell = 'null'
wb = load_workbook(sys.argv[1])
ws = wb.get_sheet_by_name('Current Items')

print("Enter search query")
query = str(raw_input(">"))

while 1 == 1:
        currentCell = ws.cell(row=rowCount, column=6).value.encode('ascii', 'replace')
        if query in currentCell.lower() or query in currentCell or query in currentCell.upper():
                found = True
                if 'OUT' in str(ws.cell(row=rowCount, column=8).value) or '0' == str(ws.cell(row=rowCount, column=10).value):
                        print currentCell + ' OUT OF STOCK'
                else:
                        print currentCell
        rowCount += 1
        if rowCount >= ws.max_row:
                if found == False:
                        print "Nothing found. You either mistyped something, or it's out of stock."
                break
